
import os
import pandas as pd
import matplotlib.pyplot as plt
# import link_distribution as ld
# import xlwt
from openpyxl import Workbook
import numpy as np

import sys
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import app.noc.link_distribution_16 as ld16

# 1, 25, 50, 75, 100, 150, 200, 500
rf_stage = {
	"lu":[97.7, 85.76, 84.24, 81.03, 78.63, 75.21, 75],
	"canneal":[96.97, 92.33, 90.07, 89.13, 83.98, 82.33, 76.34, 71.02],
	"dedup":[99, 90.27, 88.17, 81.93, 81.34, 79.34, 77.83, 72.3]
}

def validate_vector():
	file_path = "../../../../../reference/expensive_experiments/simulator/aeolus/10thJune2018/"
	file_path += "max_val_29June_actual_sim_rfk_stg_unord_nsw_0/lu_output_actual_sim_all.csv"
	act_sim_val = pd.read_csv(file_path, delim_whitespace=True, header=None).as_matrix()
	act_sim_val = act_sim_val[:,1:]
	print(act_sim_val.shape)
	print(act_sim_val)
	ld16.validate_sw_distribution(act_sim_val)
	return


def get_collated_results(BASE_FOLDER_PATH, TRAIL_LIST,
	BENCHMARK_LIST):

	benchmark_results = {}
	
	for benchmark in BENCHMARK_LIST:
		results = np.zeros((6,52))
		for idx, trail_idx in enumerate(TRAIL_LIST):
			folder_path = BASE_FOLDER_PATH+"_"+str(trail_idx)
			file_path = os.path.join(folder_path, benchmark)
			file_path += "_output_actual_sim_max.csv"
			# print(file_path)
			act_sim_val = pd.read_csv(file_path, delim_whitespace=True, header=None).as_matrix()
			# print(act_sim_val[:, 0], act_sim_val[:, 1])
			result_single = act_sim_val[:, 1]
			print(act_sim_val.shape, result_single.shape)
			results[idx,:len(result_single)] = result_single
		avg_results = np.sum(results[:5], axis=0)/(results != 0).sum(0)
		# print(np.count_nonzero(results, axis=1))
		print((results != 0).sum(0))
		# print("Average", benchmark, BASE_FOLDER_PATH)
		# print(avg_results)
		results[5,:] = avg_results
		benchmark_results[benchmark] = results

	return benchmark_results

def dump_all_algorithms(ws, offset_row, 
	offset_column, algorithm_results):
	# print(algorithm_results)

	for algorithm, algo_results in algorithm_results.items():
		# Merge all the cells and write reference
		ws.merge_cells(start_row=offset_row,start_column=offset_column,
			end_row=offset_row,end_column=offset_column+len(algo_results))
		ws.cell(row=offset_row, column=offset_column,
					value=algorithm[26:])
		offset_row+=1

		i = 0
		# for idx in [-1,0,25,50,75,100,125,150,175,200,225,\
		# 250,275,300,325,350,375,400,425,450,475,500]:
		idx_list = [-1]
		idx_list.extend(range(0,500,10))
		# print("Append ", idx_list)
		for idx in idx_list:
			# Merge all the cells and write reference
			ws.cell(row=offset_row+i+1, column=offset_column,
						value=idx)
			i+=1
		#offset_row+=1
		offset_column+=1

		i = 0
		for benchmark, benchmark_results in algo_results.items():
			# Merge all the cells and write reference
			ws.cell(row=offset_row, column=offset_column+i,
						value=benchmark)
			i+=1
		offset_row+=1

		

		i = 0
		for benchmark, benchmark_results in algo_results.items():
			for j in range(len(benchmark_results[5,:])):

				# Merge all the cells and write reference
				ws.cell(row=offset_row+j, column=offset_column+i,
							value=benchmark_results[5,j])
			i+=1

		offset_row+= len(benchmark_results[5,:]) + 1
#			print(algorithm, benchmark, benchmark_results[5,:])
		offset_column-=1
	return offset_row, offset_column

def write_collated_xlsx(algorithm_results, max_val_folder, xlsx_filename):
	wb = Workbook()
	ws = wb.create_sheet("EDP sheet", 0)
	offset_row = 1
	offset_column = 3
	offset_row, offset_column = \
		dump_all_algorithms(ws, offset_row, offset_column, algorithm_results)

	xlsx_file_path = os.path.join(max_val_folder, xlsx_filename)
	print(xlsx_file_path)
	wb.save(xlsx_filename)
	return


validate_vector()

BASE_PATH = "../../../../../../reference/expensive_experiments/simulator/aeolus/12thSept/"
BASE_PATH = "../../../../../../reference/expensive_experiments/simulator/aeolus/10Jan2018/"
BASE_PATH = "../../../../../reference/expensive_experiments/simulator/aeolus/20thJune2018/"
BASE_PATH = "../../../../../reference/expensive_experiments/simulator/aeolus/10thJune2018/"

FOLDER_LIST = [
"max_val_15October_actual_sim_rf_rls_ord",
"max_val_15October_actual_sim_rf_stg_ord",
#"max_val_13September_actual_sim_rf_rls_unord",
#"max_val_13September_actual_sim_rf_stg_unord"
]
FOLDER_LIST = [
"max_val_17October_actual_sim_rf_rls_ord",
"max_val_17October_actual_sim_rf_stg_ord",
]
# FOLDER_LIST = [
# "max_val_12September_actual_sim_rf_rls_ord",
# "max_val_12September_actual_sim_rf_stg_ord",
# # "max_val_13September_actual_sim_rf_rls_unord",
# # "max_val_13September_actual_sim_rf_stg_unord"
# ]
FOLDER_LIST = [
"max_val_23October_actual_sim_rf_rls_ord",
"max_val_23October_actual_sim_rf_stg_ord",
]
FOLDER_LIST = [
"max_val_17November_actual_sim_rbf_stg_ord",
"max_val_17November_actual_sim_rf_stg_ord"

]
FOLDER_LIST = [
"max_val_26December_actual_sim_rf_stg_ord",
"max_val_26December_actual_sim_rbf_stg_ord"
]
FOLDER_LIST = [
"max_val_21June_actual_sim_rfk_stg_unord",
"max_val_21June_actual_sim_fck_stg_unord",
"max_val_21June_actual_sim_rfk_stg_unord_nsw",
"max_val_21June_actual_sim_fck_stg_unord_nsw"
]
FOLDER_LIST = [
"max_val_29June_actual_sim_rfk_stg_unord_nsw",
"max_val_29June_actual_sim_fck_stg_unord_nsw",
"max_val_29June_actual_sim_rfk_smac_unord_nsw",
"max_val_29June_actual_sim_fck_smac_unord_nsw",
"max_val_29June_actual_sim_rfk_smac_unord",
"max_val_29June_actual_sim_rfk_smac_unord",
"max_val_29June_actual_sim_rfk_stg_unord",
"max_val_29June_actual_sim_fck_stg_unord"
]


TRAIL_IDX_LIST = [
[0, 1, 2, 4, 5],
[6, 8, 9, 10, 11],
# [0, 1, 2, 4, 5],
#[6, 8, 9, 10, 11],
]
TRAIL_IDX_LIST = [
[0],
[1],
[2],
[3],
[4],
[5],
[6],
[7]
]

BENCHMARK_LIST = ["canneal", "dedup",  "lu"]#, "fluid", "radix"] #  "fft", 

algorithm_results = {}
for idx, folder in enumerate(FOLDER_LIST):
	print(idx, folder)
	algorithm_results[folder] = \
		get_collated_results(os.path.join(BASE_PATH, folder), 
			TRAIL_IDX_LIST[idx], BENCHMARK_LIST)

write_collated_xlsx(algorithm_results, "./", "collated_testing.xlsx")
