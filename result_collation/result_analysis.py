import os
import pandas as pd
import matplotlib.pyplot as plt
import link_distribution as ld
# import xlwt
from openpyxl import Workbook


reference_3d_mesh = {
	"canneal":{ "latency": 17.194, "energy": 2.18E-09,"edp": 3.74E-08},
	"dedup":{ "latency": 16.926, "energy": 2.15E-09, "edp": 3.63E-08},
	"fft":{ "latency": 18.227, "energy": 2.24E-09, "edp": 4.08E-08},
	"fluid":{"latency": 18.184, "energy": 2.24E-09, "edp": 4.07E-08},
	"lu":{"latency": 17.269, "energy": 2.17E-09, "edp": 3.75E-08},
	"radix":{"latency": 18.17, "energy": 2.22E-09, "edp": 4.04E-08},
	"vips":{"latency":  17.513, "energy": 2.19E-09, "edp": 3.84E-08},
	"water":{"latency":  15.1, "energy": 1.97E-09, "edp": 2.98E-08}
}

# print(reference_3d_mesh["water"]["edp"])

def plot_all_benchmark(max_val_folder, benchmark_list, 
	act_edp_ratio_list, act_sim_val_list):

	idx = 0
	for benchmark in benchmark_list:
		f = plt.figure()
		plt.plot(act_sim_val_list[idx][:,0], act_edp_ratio,  "o")
		plt.xlabel('Num Iterations (N)')
		plt.ylabel('EDP Ratio')
		plt.title('Energy Delay Product '+benchmark)
		# plt.show()
		plot_figure_name = os.path.join(max_val_folder, benchmark+"_image.png")
		f.savefig(plot_figure_name, bbox_inches='tight')
		idx = idx + 1

	return

def dump_ref(ws, offset_row, offset_column, benchmark_list):
	# offset_row = 1
	# offset_column = 1

	# Merge all the cells and write reference
	ws.merge_cells(start_row=offset_row,start_column=offset_column,
		end_row=offset_row,end_column=offset_column+len(benchmark_list))
	ws.cell(row=offset_row, column=offset_column,
				value="Reference")
	offset_row+=1

	# Write the table heading
	i = 0
	for benchmark in benchmark_list:
		ws.cell(row=offset_row, column=offset_column+i+1,
				value=benchmark)
		i+=1
	offset_row+=1

	# Write the key values
	i = 0
	for key, value in reference_3d_mesh[benchmark_list[0]].iteritems():
		ws.cell(row=offset_row+i, column=offset_column,
				value=key)
		i+=1
	offset_column+=1

	# Dump the reference values
	i = 0
	for benchmark in benchmark_list:
		j = 0
		for key, value in reference_3d_mesh[benchmark].iteritems():
			ws.cell(row=offset_row+j, column=offset_column+i,
				value=value)
			j = j + 1
		i = i + 1

	offset_row+=i
	offset_column-=1

	return offset_row, offset_column

def dump_edp_benchmark(ws, offset_row, offset_column, 
	benchmark, act_sim_val, act_edp_ratio):

	# Merge all the cells and write reference
	ws.merge_cells(start_row=offset_row,start_column=offset_column,
		end_row=offset_row,end_column=offset_column+4)
	ws.cell(row=offset_row, column=offset_column,
				value=benchmark)
	offset_row+=1

	# Headings: num_iter, latency, energy, edp,edp ratio
	ws.cell(row=offset_row, column=offset_column,
				value="num iter")
	ws.cell(row=offset_row, column=offset_column+1,
				value="latency")
	ws.cell(row=offset_row, column=offset_column+2,
				value="energy")
	ws.cell(row=offset_row, column=offset_column+3,
				value="edp")
	ws.cell(row=offset_row, column=offset_column+4,
				value="edp_ratio")
	offset_row+=1

	# num iter
	for i in range(len(act_sim_val)):
		ws.cell(row=offset_row+i, column=offset_column,
				value=act_sim_val[i,0])
		ws.cell(row=offset_row+i, column=offset_column+1,
				value=act_sim_val[i,2])
		ws.cell(row=offset_row+i, column=offset_column+2,
				value=act_sim_val[i,3])
		ws.cell(row=offset_row+i, column=offset_column+3,
				value=act_sim_val[i,4])
		ws.cell(row=offset_row+i, column=offset_column+4,
				value=act_edp_ratio[i])
	offset_row+=i

	
	return offset_row, offset_column



def write_xlsx(max_val_folder, benchmark_list, 
	act_edp_ratio_list, act_sim_val_list,
	xlsx_filename):
	wb = Workbook()
	ws = wb.create_sheet("EDP sheet", 0)
	offset_row = 1
	offset_column = 1
	offset_row, offset_column = \
		dump_ref(ws, offset_row, offset_column, benchmark_list)

	idx = 0
	for benchmark in benchmark_list:
		offset_row, offset_column = \
			dump_edp_benchmark(ws, offset_row, offset_column, 
				benchmark, act_sim_val_list[idx], act_edp_ratio_list[idx])
		idx = idx + 1

	xlsx_file_path = os.path.join(max_val_folder, xlsx_filename)
	print(xlsx_file_path)
	wb.save(xlsx_filename)
	return

def get_edp_ratio_list(max_val_folder, benchmark_list):
	act_edp_ratio_list = []
	act_sim_val_list = []
	for benchmark in benchmark_list:
		filename = benchmark+"_actual_sim_max.csv"
		max_val_file_name = os.path.join(max_val_folder, filename)
		print(max_val_file_name)


		# 0-num_iterations, 1-dummy_simulator_val
		# 2-4 - Actual sim val
		# 2-latency, 3-energy, 4-EDP
		act_sim_val = pd.read_csv(max_val_file_name, delim_whitespace=True, header=None).as_matrix()
		# print(act_sim_val[:,1])
		# print(act_sim_val[:,2])
		# print(act_sim_val)

		# Subtract 6 from latency since each simulator call involves 6 packets
		# Reference values
		ref_latency = reference_3d_mesh[benchmark]["latency"] - 6
		ref_edp = reference_3d_mesh[benchmark]["edp"]

		# Latency Ratio
		act_latency = act_sim_val[:,2] - 6
		act_latency_ratio = act_latency/ref_latency
		act_energy_ratio = act_latency_ratio*act_sim_val[:,3]
		act_edp = act_sim_val[:,2]*act_energy_ratio
		act_edp_ratio = act_edp/ref_edp
		# print(act_latency)
		# print(act_latency_ratio)
		# print(act_energy_ratio)
		# print(act_edp)
		print(act_edp_ratio)
		print(act_sim_val[:,0])
		act_edp_ratio_list.append(act_edp_ratio)
		act_sim_val_list.append(act_sim_val)


	return act_edp_ratio_list, act_sim_val_list

def dump_benchmarks(max_val_folder, benchmark_list, xlsx_filename):
	act_edp_ratio_list, act_sim_val_list = \
		get_edp_ratio_list(max_val_folder, benchmark_list)
	# plot_all_benchmark(max_val_folder, benchmark_list, act_edp_ratio_list, act_sim_val_list)
	write_xlsx(max_val_folder, benchmark_list, 
		act_edp_ratio_list, act_sim_val_list, xlsx_filename)
	return


def validate_graph(max_val_folder, benchmark_list):
	for benchmark in benchmark_list:
		filename = benchmark+"_output_dummy_sim_max.csv"
		max_val_file_name = os.path.join(max_val_folder, filename)
		print(max_val_file_name)
		act_sim_val = pd.read_csv(max_val_file_name, delim_whitespace=True, header=None).as_matrix()
		print(act_sim_val.shape)
		for i in range(act_sim_val.shape[0]):
			feature_vector = act_sim_val[i,2:]
			ld.LinkDistribution().check_sw_connectivity(feature_vector)


BASE_PATH = "../../../../../reference/expensive_experiments/simulator/aeolus/12thSept/"
FOLDER_LIST = [
# "max_val_06August_actual_sim_wl_rls_ord",
"max_val_06August_actual_sim_wl_stg_ord",
# "max_val_30July_actual_sim_stg_ord",
# "max_val_31July_actual_sim_stg_unord"
]

BENCHMARK_LIST = ["canneal", "dedup",  "fft",  "radix", "vips", "water"]
# BENCHMARK_LIST = ["canneal", "dedup",  "lu"]

# act_edp_ratio_list = plot_all_benchmark(MAX_VAL_FOLDER_1, BENCHMARK_LIST)
# plot_graph(act_edp_ratio_list)
# act_edp_ratio_list = plot_all_benchmark(MAX_VAL_FOLDER_2, BENCHMARK_LIST)
# plot_graph(act_edp_ratio_list)

for folder in FOLDER_LIST:
	dump_benchmarks(os.path.join(BASE_PATH, folder), 
		BENCHMARK_LIST, folder+".xlsx")
# dump_benchmarks(MAX_VAL_FOLDER_2, BENCHMARK_LIST)
# dump_benchmarks(MAX_VAL_FOLDER_3, BENCHMARK_LIST)
# dump_benchmarks(MAX_VAL_FOLDER_4, BENCHMARK_LIST)

# dump_xls()
# dump_xlsx()

# validate_graph(MAX_VAL_FOLDER_1, BENCHMARK_LIST)
# validate_graph(MAX_VAL_FOLDER_2, BENCHMARK_LIST)
# validate_graph(MAX_VAL_FOLDER_3, BENCHMARK_LIST)
# validate_graph(MAX_VAL_FOLDER_4, BENCHMARK_LIST)


def plot_graph(benchmark_list):
	print(benchmark_list[0])
	t = range(len(benchmark_list[0]))
	for benchmark in benchmark_list:
		plt.plot(t, benchmark)
	plt.xlabel('time (s)')
	plt.ylabel('voltage (mV)')
	plt.title('About as simple as it gets, folks')
	plt.grid(True)
	plt.savefig("test.png")
	plt.show()

def dump_xls():
	style0 = xlwt.easyxf('font: name Times New Roman, color-index red, bold on',
	    num_format_str='#,##0.00')
	style1 = xlwt.easyxf(num_format_str='D-MMM-YY')

	wb = xlwt.Workbook()
	ws = wb.add_sheet('A Test Sheet')

	ws.write(0, 0, 1234.56, style0)
	ws.write(1, 0, datetime.now(), style1)
	ws.write(2, 0, 1)
	ws.write(2, 1, 1)
	ws.write(2, 2, xlwt.Formula("A3+B3"))

	wb.save('example.xlsx')

def dump_xlsx():
	wb = Workbook()
	
	wb._active_sheet_index = 0 # xetting the sheet
	ws = wb.active
	ws.title = "New Title"

	ws = wb.create_sheet("EDP sheet", 0)

	ws['A4'] = 4
	ws.cell(row=4, column=2, value=10)

	wb.save('example.xlsx')


