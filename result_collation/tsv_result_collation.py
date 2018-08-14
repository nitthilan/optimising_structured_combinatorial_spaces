import os
from openpyxl import Workbook
import numpy as np



def dump_results_sheet(ws, offset_row, offset_column, benchmark_results):
	for tsv_size, tsv_size_results in benchmark_results.iteritems():
		# Merge all the cells and write reference
		# ws.merge_cells(start_row=offset_row,start_column=offset_column,
		# 	end_row=offset_row,end_column=offset_column+len(algo_results))
		ws.cell(row=offset_row, column=offset_column,
					value=tsv_size)
		# offset_row+=1

		i = 0
		for value in tsv_size_results:
			# Merge all the cells and write reference
			ws.cell(row=offset_row+i+2, column=offset_column,
						value=value)
			i+=1
		#offset_row+=1
		offset_column+=1
	return offset_row, offset_column

def write_collated_xlsx(algorithm_results, max_val_folder, xlsx_filename):
	wb = Workbook()
	i = 0
	for benchmark, benchmark_results in algorithm_results.iteritems():
		print(benchmark, benchmark_results)

		ws = wb.create_sheet(benchmark, i)
		offset_row = 1
		offset_column = 3
		offset_row, offset_column = \
			dump_results_sheet(ws, offset_row, offset_column, benchmark_results)
		i += 1

	xlsx_file_path = os.path.join(max_val_folder, xlsx_filename)
	print(xlsx_file_path)
	wb.save(xlsx_filename)
	return

def parse_tsv_time_values(base_path, benchmark_list, 
	tsv_size_list, reference_time_values):
	results = {}
	design_results = {}
	for benchmark in benchmark_list:
		results[benchmark] = {}
		design_results[benchmark] = {}
		for tsv_size in tsv_size_list:
			basefolder_path = os.path.join(base_path, benchmark)
			#basefolder_path = basefolder_path+"_"+str(tsv_size)
			basefolder_path = os.path.join(basefolder_path, str(tsv_size))
			filepath = os.path.join(basefolder_path, "modelTrainDataStage.txt")
			
			if(not os.path.isfile(filepath)):
				print("NOT A VALID FILE PATH ", filepath)
				continue
			with open(filepath) as f:
			    content = f.readlines()
			values = []
			design_val = []
			max_val = 0
			actual_val = 0
			reference_val = 0
			for x in content[3:]:
				split_x = x.strip().split()
				# print(len(split_x))
				if(len(split_x) == 384):
					design_split = x.strip()[1:-1].split(",")
					design_val = [int(x.strip()) for x in design_split]

				if(len(split_x) > 1 and len(split_x) < 20):
					#print(split_x)
					ratio_val = float(split_x[-1])/reference_time_values[benchmark]
					values.append(ratio_val)
					if(ratio_val > max_val):
						max_val = ratio_val
						actual_val = split_x[-1]
						reference_val = reference_time_values[benchmark]
						design_val = split_x
			# print(benchmark, tsv_size, len(values))
			# print(values)
			results[benchmark][tsv_size] = values
			design_results[benchmark][tsv_size] = {
				"max_val" : max_val,
				"actual_val" : actual_val,
				"reference_val" : reference_val,
				"design_val" : design_val
			}
	# print(design_results)
	return results, design_results

def get_greedy_ref_design(basepath, benchmark_list, tsv_size_list):
	greedy_design_result = {}
	for benchmark in benchmark_list:
		greedy_design_result[benchmark] = {}
		for tsv in tsv_size_list:
			filepath = os.path.join(basepath, benchmark)
			filepath = os.path.join(filepath, str(tsv)+".txt")
			print(filepath)
			with open(filepath, 'r') as f:
				content = f.readlines()
			
			goodSol = list()

			for i in range(48):
				for j in range(8):
					goodSol.append(int(content[i].split('\n')[0].split('\t')[j]))
			greedy_design_result[benchmark][tsv] = goodSol

	# print(greedy_design_result)
	return greedy_design_result

# BASE_PATH = "../../../../../../reference/expensive_experiments/simulator_tsv/lab_machine/output_tsv_results/merged"
# BASE_PATH = "../../../../../../reference/expensive_experiments/simulator_tsv/lab_machine/output_tsv_results_6thOct/merged"
# BASE_PATH = "../../../../../../reference/expensive_experiments/simulator_tsv/lab_machine/output_tsv_results_10thOct/merged"
# BASE_PATH = "../../../../../../reference/expensive_experiments/simulator_tsv/lab_machine/output_tsv_results_16thOct/merged"
# BASE_PATH = "../../../../../../reference/expensive_experiments/simulator_tsv/lab_machine/output_tsv_results_23rdOct/merged"
# BASE_PATH = "../../../../../../reference/expensive_experiments/simulator_tsv/lab_machine/output_tsv_results_31stOct/merged"
BASE_PATH = "../../../../../../reference/expensive_experiments/simulator_tsv/lab_machine/250K_runs_1st_dec/"


BENCHMARK_LIST = [
"Canneal",
"Dedup",
"Vips"
]
TSV_SIZE_LIST = [9, 19, 28, 38, 57]



REFERENCE_TIME_VALUES = {
	"Canneal": 4*63528.4,
	"Dedup": 4*373570, 
	"fft": 4*267211, 
	"radix":4*537988,
	"Vips":4*1098150, 
	"water":4*781459,
}

algorithm_results, design_results = \
	parse_tsv_time_values(BASE_PATH, BENCHMARK_LIST, 
		TSV_SIZE_LIST, REFERENCE_TIME_VALUES)
write_collated_xlsx(algorithm_results, "./", "tsv_results.xlsx")

# GREEDY_PATH = "../arnab/greedy"
# GREEDY_BENCHMARK = [
# "canneal",
# "dedup",
# "vips"
# ]
# GREEDY_TSV_LIST = [9, 19, 38, 57]
# greedy_design_result = get_greedy_ref_design(GREEDY_PATH, GREEDY_BENCHMARK, GREEDY_TSV_LIST)

# for benchmark in BENCHMARK_LIST:
# 	for tsv in GREEDY_TSV_LIST:
# 		small_case = benchmark.lower()
# 		greedy_design = np.array(greedy_design_result[small_case][tsv])
# 		current_max = np.array(design_results[benchmark][tsv]["design_val"])
# 		diff_value = greedy_design - current_max
# 		# print(diff_value)
# 		print(benchmark, np.sum(diff_value), np.sum(np.abs(diff_value)), tsv)
# 		# print(np.mod(np.where(diff_value>0), 8), np.mod(np.where(diff_value<0), 8))
# 		layer_idx_greedy = np.divide(np.where(diff_value>0),8)
# 		layer_idx_current = np.divide(np.where(diff_value<0), 8)
# 		len_mid_greedy = len(layer_idx_greedy[(layer_idx_greedy >= 16) & (layer_idx_greedy < 32)])
# 		len_mid_current = len(layer_idx_current[(layer_idx_current >= 16) & (layer_idx_current < 32)])
# 		# print(layer_idx_greedy, layer_idx_current)
# 		print(len_mid_greedy, len(layer_idx_greedy[0]), len_mid_current, len(layer_idx_current[0]))




# ../../../../../simulator/simulator_tsv_2\Dedup\38\modelTrainDataStage.txt
# ('Dedup', 38, 16)
# [1.2887276815590116, 1.3141499585084455, 1.3765532564178065, 1.4236475091682952, 1.478386915437535, 1.4786974328773723, 1.4526514441737828, 1.2536445646063656, 1.44688010279198, 1.4089675295125412, 1.
# 440495757153947, 1.2306823353052976, 1.4057097732687314, 1.4418074256498112, 1.4773456112642878, 1.1815750729448296]
# ../../../../../simulator/simulator_tsv_2\Dedup\57\modelTrainDataStage.txt
# ('Dedup', 57, 21)
# [1.7041063254543996, 1.3065690499772467, 1.3216692989265733, 1.5294188505500976, 1.444899215675777, 1.6133522499130015, 1.6677168937548519, 1.4826993602269989, 1.6398104772867201, 1.8558931391707043,
# 1.677091308188559, 1.8586422892630565, 1.6621302567122627, 1.6176298953342079, 1.4970848836898039, 1.6684878336054823, 1.617819953422384, 1.6084348314907515, 1.7221270444628851, 1.752766549776481, 1.4
# 052386433600128]
# ../../../../../simulator/simulator_tsv_2\Vips\9\modelTrainDataStage.txt
# ('Vips', 9, 34)
# [1.068196512316168, 1.1329326594727496, 1.023393889723626, 1.1036015116332012, 1.1329326594727496, 1.1329326594727496, 1.1453990802713654, 1.1453990802713654, 1.1453990802713654, 1.194490734416974, 1.
# 2258252515594408, 1.179620270454856, 1.1681828529800118, 1.1621636388471521, 1.2893047397896462, 1.1022264717934709, 1.2734599098483814, 1.104266265992806, 1.1765970040522697, 1.2734599098483814, 1.25
# 80977097846378, 1.2580977097846378, 1.216263716250057, 1.2583253653872422, 1.2893047397896462, 1.240222191868142, 1.2258252515594408, 1.257214406046533, 1.2258252515594408, 1.2601557164321815, 1.28930
# 47397896462, 1.2700359695852115, 1.2734599098483814, 1.2258252515594408]
# ../../../../../simulator/simulator_tsv_2\Vips\19\modelTrainDataStage.txt
# ('Vips', 19, 31)
# [1.0689432226927105, 1.144224377361927, 1.190019578381824, 1.2856895688202887, 1.2056549651686927, 1.3682101716523243, 1.388298502026135, 1.3265765150480353, 1.3145380867823158, 1.3375586213176707, 1.
# 468970541365023, 1.1925784273550972, 1.3331785275235624, 1.3078723307380595, 1.2944679688567136, 1.4880936119837909, 1.2441105495606246, 1.4011564904612301, 1.277657879160406, 1.5048217456631607, 1.51
# 50116104357327, 1.4552929927605518, 1.4327824067750308, 1.3605154122842964, 1.329262851158767, 1.4164640531803487, 1.428220188498839, 1.3415471474752994, 1.4397395619906206, 1.4664116923917498, 1.4926
# 922551563995]
# ../../../../../simulator/simulator_tsv_2\Vips\28\modelTrainDataStage.txt
# ('Vips', 28, 18)
# [1.20270454855894, 1.2411874516231844, 1.313572827027273, 1.3893183991258025, 1.3944269908482447, 1.4020580066475437, 1.2224195237444793, 1.393343350179848, 1.3388517051404636, 1.3567090106087512, 1.4
# 637890998497474, 1.4326002822929473, 1.3825524746164004, 1.3718890861904112, 1.411227974320448, 1.4440650184401038, 1.4760825023903839, 1.4108819378044894]
# ../../../../../simulator/simulator_tsv_2\Vips\38\modelTrainDataStage.txt
# ('Vips', 38, 24)
# [1.2832217820880572, 1.1945180530892865, 1.233365205117698, 1.485134089149934, 1.542667213040113, 1.4343851022173655, 1.5138551199745025, 1.7767882347584574, 1.277138824386468, 1.6216454946956245, 1.5
# 006602012475527, 1.445112234212084, 1.467768519783272, 1.427446159449984, 1.5859308837590493, 1.2182579793288713, 1.442252879843373, 1.608568956882029, 1.4939124891863589, 1.5270136138050356, 1.563875
# 608978737, 1.568756545098575, 1.396931202476893, 1.368556208168283]
# ../../../../../simulator/simulator_tsv_2\Vips\57\modelTrainDataStage.txt
# ('Vips', 57, 20)
# [1.457851841733825, 1.2997677912853436, 1.4016482265628558, 1.7186996311979237, 1.4687519919865228, 1.5062058917269954, 1.3668715567090106, 1.4601739288803897, 1.6617037745298913, 1.707426125756955, 1
# .468870372899877, 1.5842098074033601, 1.760397031370942, 1.7022628966898876, 1.3811501161043573, 1.7857851841733825, 1.8049173610162546, 1.937066885216045, 1.5671265309839275, 1.6495014342302965]